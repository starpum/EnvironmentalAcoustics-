# LD processing 
# Processes files from Larson Davis soundmeters into graphs and synthetizes the results into a single excel workbook

import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl import Workbook
import os
import numpy as np
import matplotlib as plt

# main script 
def main():
    root = tk.Tk()
    root.withdraw()

    # Select folder with msm datas
    folder = filedialog.askdirectory()

    ## Variables creation
    n = 0
    f_11 = ["Hz", 8, 16, 31.5, 63, 125, 250, 500, 1000, 2000, 4000, 8000, 16000]
    f_13 = ["Hz", 6.3, 8.0, 10.0, 12.5, 16.0, 20.0, 25.0, 31.5, 40.0, 50.0, 63.0, 80.0, 100, 125, 160, 200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500, 3150, 4000, 5000, 6300, 8000, 10000, 12500, 16000, 20000]
    f_head = [6.3, 8.0, 10.0, 12.5, 16.0, 20.0, 25.0, 31.5, 40.0, 50.0, 63.0, 80.0, 100, 125, 160, 200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500, 3150, 4000, 5000, 6300, 8000, 10000, 12500, 16000, 20000]

    #Output workbook 
    wb_out = Workbook()         
    ws_out = wb_out.active
    ws_out.title = 'Summary'
    headers = ["Point name", "Date&Time","LAeq", "LA50", "LA90", "LA95","Lf"]
    ws_out.append(headers)
    headers = ["", "", "dB(A)", "dB(A)", "dB(A)", "dB(A)"] + f_head
    ws_out.append(headers)

    #grab list of files in folder
    xls_files = [f for f in os.listdir(folder)
                if f.lower().endswith(".xlsx") and os.path.isfile(os.path.join(folder, f))]
    pt_name = [""] * len(xls_files) #Point name for saving

    ## Processing of datas
    for f in os.listdir(folder):
        #Open the worksheet
        path = os.path.join(folder,f)
        wb = openpyxl.load_workbook(path)
        
        #Ask for point name 
        print(f"Measurement point name for '{f}':")
        pt_name[n] = input("> ")
        
        # Get relevant overall datas from the "Summary" sheet
        ws = wb["Summary"]

        pt = pt_name[n]
        Date = ws['B14'].value
        LAeq = ws['B73'].value
        LA50 = ws['B91'].value
        LA90 = ws['B92'].value
        LA95 = ws['B93'].value
        
        ws = wb["OBA"]
        s_13 = [ws.cell(row= 9,column=i).value for i in range(2,38)]
        output_data = [pt, Date, LAeq, LA50, LA90, LA95] + s_13
        ws_out.append(output_data)
        
        n = n+1
        
        wb_out.save('Results.xlsx')
        
        
# Code execution 
if __name__ == "__main__":

    main()